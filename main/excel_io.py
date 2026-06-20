import shutil
from copy import copy

from openpyxl import load_workbook

from config import BACKUP_PATH, COLUMNS, EXCEL_PATH, LOCKED_COLUMNS, SHEET_NAME


def backup_excel() -> None:
    if EXCEL_PATH.exists():
        shutil.copy2(EXCEL_PATH, BACKUP_PATH)


def load_contact_list() -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return COLUMNS, []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        values = [str(v).strip() if v is not None else "" for v in row]
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        record = dict(zip(headers, values))
        if any(record.get(col, "") for col in COLUMNS[:10]):
            record["_row"] = len(records) + 2
            records.append(record)
    return headers, records


def find_row_by_company(company_name: str) -> int | None:
    from dedup import similarity

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    best_idx = None
    best_score = 0.0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        score = similarity(company_name, name)
        if score > best_score:
            best_score = score
            best_idx = idx
    wb.close()
    return best_idx if best_score >= 0.8 else None


def _header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            mapping[str(cell.value).strip()] = col_idx
    return mapping


def enrich_row(
    company_name: str,
    updates: dict[str, str],
    notes_append: str = "",
    allow_overwrite: set[str] | None = None,
) -> list[str]:
    allow_overwrite = allow_overwrite or set()
    backup_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    headers = _header_map(ws)
    row_idx = find_row_by_company(company_name)
    if row_idx is None:
        wb.close()
        raise ValueError(f"Company not found: {company_name}")

    changed: list[str] = []
    for col, new_val in updates.items():
        if col not in headers:
            continue
        if col in LOCKED_COLUMNS:
            raise ValueError(f"Attempted to modify locked column: {col}")
        col_idx = headers[col]
        current = ws.cell(row=row_idx, column=col_idx).value
        current_str = str(current).strip() if current is not None else ""
        if current_str and col not in allow_overwrite:
            continue
        ws.cell(row=row_idx, column=col_idx, value=new_val)
        changed.append(col)

    if notes_append and "Notes" in headers:
        notes_idx = headers["Notes"]
        existing = ws.cell(row=row_idx, column=notes_idx).value
        existing_str = str(existing).strip() if existing is not None else ""
        combined = f"{notes_append}\n{existing_str}".strip() if existing_str else notes_append
        ws.cell(row=row_idx, column=notes_idx, value=combined)
        if notes_append:
            changed.append("Notes")

    wb.save(EXCEL_PATH)
    wb.close()
    return changed


def append_row(record: dict[str, str], notes_append: str = "") -> int:
    backup_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    headers = _header_map(ws)
    new_row = ws.max_row + 1

    for col in LOCKED_COLUMNS:
        if col in record and record[col]:
            raise ValueError(f"Cannot set locked column on append: {col}")

    for col in COLUMNS:
        if col not in headers:
            continue
        val = record.get(col, "")
        if col == "Notes" and notes_append:
            val = notes_append if not val else f"{notes_append}\n{val}"
        if val:
            ws.cell(row=new_row, column=headers[col], value=val)

    wb.save(EXCEL_PATH)
    wb.close()
    return new_row
